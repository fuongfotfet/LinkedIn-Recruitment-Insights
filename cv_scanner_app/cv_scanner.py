import streamlit as st
import google.generativeai as genai
import os
import PyPDF2 as pdf
from dotenv import load_dotenv
import json
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

# Load environment variables
load_dotenv()

# Configure Streamlit page settings
st.set_page_config(
    page_title="Smart ATS",
    page_icon="👨‍💼",
    layout="centered",
)

# Sidebar to input Google API Key
st.sidebar.title("Smart ATS Configuration")
API_KEY = st.sidebar.text_input("Enter your Google API Key", type="password")
st.sidebar.subheader("Don't have a Google API Key?")
st.sidebar.write("Visit [Google Makersuite](https://makersuite.google.com/app/apikey) and log in with your Google account. Then click on 'Create API Key'.")

# Check if API key is provided
if not API_KEY:
    st.error("Please enter your Google API Key.")
    st.stop()

# Function to configure Gemini AI model with the provided API key
def configure_gemini_api(api_key):
    genai.configure(api_key=api_key)

# Configure Gemini AI model with the provided API key
configure_gemini_api(API_KEY)

# Function to get response from Gemini AI
def get_gemini_response(input):
    model = genai.GenerativeModel('gemini-pro')
    response = model.generate_content(input)
    return response.text

# Function to extract text from uploaded PDF file
def input_pdf_text(uploaded_file):
    reader = pdf.PdfReader(uploaded_file)
    text = ""
    for page in range(len(reader.pages)):
        text += reader.pages[page].extract_text() or ""
    return text

# Function to save CV evaluation to Excel and store CV files
def save_to_excel_and_store_files(cv_data_list):
    file_path = "cv_evaluations.xlsx"
    if not os.path.exists("uploaded_cvs"):
        os.makedirs("uploaded_cvs")
    
    for cv_data in cv_data_list:
        # Save the uploaded CV file
        file_name = cv_data["File Name"]
        file_content = cv_data.pop("File Content")
        with open(f"uploaded_cvs/{file_name}", "wb") as f:
            f.write(file_content)

    new_data = pd.DataFrame(cv_data_list)
    if os.path.exists(file_path):
        with pd.ExcelWriter(file_path, mode='a', engine='openpyxl', if_sheet_exists='overlay') as writer:
            df = pd.read_excel(file_path)
            new_data.to_excel(writer, index=False, header=False, startrow=len(df) + 1)
            workbook = writer.book
            worksheet = writer.sheets['Sheet1']
            for index, link in enumerate(new_data["File Path"], start=len(df) + 2):
                cell = f'B{index}'
                worksheet[cell].hyperlink = link
                worksheet[cell].font = Font(color="0000FF", underline="single")
    else:
        new_data.to_excel(file_path, index=False)
        workbook = load_workbook(file_path)
        worksheet = workbook.active
        for index, link in enumerate(new_data["File Path"], start=2):
            cell = f'B{index}'
            worksheet[cell].hyperlink = link
            worksheet[cell].font = Font(color="0000FF", underline="single")
        workbook.save(file_path)

# Prompt Template
input_prompt = """
Act Like a skilled or very experienced ATS (Application Tracking System)
with a deep understanding of the tech field, software engineering, data science, data analyst
and big data engineering. Your task is to evaluate the resume based on the given job description for Business Intelligence Analyst. This company is the biggest bank in Vietnam.
You must consider the job market is very competitive and you should provide the 
best assistance for improving the resumes. Assign the percentage Matching based 
on JD and the missing keywords with high accuracy.
resume:{text}
description:{jd}

I want the response in one single string having the structure
{{"JD Match":"%","MissingKeywords":[],"Profile Summary":""}}
"""

# Streamlit app
st.title("Resume Matcher ATS")
st.markdown("Made by 😎 [Hardik](https://www.linkedin.com/in/hardikjp/)")
st.markdown("Modified by [fuong](https://www.linkedin.com/in/fuongfotfet/)")
jd = st.text_area("Paste the Job Description")
uploaded_files = st.file_uploader("Upload Your Resumes", type="pdf", accept_multiple_files=True, help="Please upload the PDFs")
submit = st.button("Submit")

if submit:
    if uploaded_files:
        cv_data_list = []
        for uploaded_file in uploaded_files:
            text = input_pdf_text(uploaded_file)
            response = get_gemini_response(input_prompt.format(text=text, jd=jd))
            st.subheader(f"Response for {uploaded_file.name}:")
            try:
                parsed_response = json.loads(response)
                for key, value in parsed_response.items():
                    st.write(f"**{key}:** {value}")

                cv_data = {
                    "File Name": uploaded_file.name,
                    "File Path": os.path.abspath(f"uploaded_cvs/{uploaded_file.name}"),
                    "Job Description": jd,
                    "JD Match": parsed_response["JD Match"],
                    "Missing Keywords": ", ".join(parsed_response["MissingKeywords"]),
                    "Profile Summary": parsed_response["Profile Summary"],
                    "File Content": uploaded_file.getvalue()
                }
                cv_data_list.append(cv_data)
            except json.JSONDecodeError:
                st.error(f"Failed to parse response from the AI model for {uploaded_file.name}. Please try again.")
        
        if cv_data_list:
            save_to_excel_and_store_files(cv_data_list)
            st.success("The resume evaluations were saved to cv_evaluations.xlsx")